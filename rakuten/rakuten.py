import openpyxl
from gui import *
from number_parser import *
from openpyxl.utils import *
from openpyxl.styles import *

def set_auto_column_widths(sheet):
    '''
    This function sets the column widths automatically
    '''
    column_widths = {}
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value:
                # add 4 at the end because Chinese chars take larger spaces
                column_widths[cell.column] = max((column_widths.get(cell.column, 0), len(cell.value)+4)) 
    for i, column_width in column_widths.items():
        sheet.column_dimensions[i].width = column_width

    sheet.column_dimensions[get_column_letter(7)].width = 13
    sheet.column_dimensions[get_column_letter(8)].width = 13
    sheet.column_dimensions[get_column_letter(9)].width = 40
    sheet.column_dimensions[get_column_letter(10)].width = 20
    return

def set_auto_column_widths_F(sheet):
    '''
    This function sets the column widths automatically
    '''
    column_widths = {}
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value:
                # add 4 at the end because Chinese chars take larger spaces
                column_widths[cell.column] = max((column_widths.get(cell.column, 0), len(cell.value)+4)) 
    for i, column_width in column_widths.items():
        sheet.column_dimensions[i].width = column_width

    sheet.column_dimensions[get_column_letter(7)].width = 13
    sheet.column_dimensions[get_column_letter(8)].width = 13
    sheet.column_dimensions[get_column_letter(9)].width = 40
    sheet.column_dimensions[get_column_letter(11)].width = 40

#   set row height
#    for col in range(5, sheet.max_row-1):
#        sheet.row_dimensions[col].height = 30
    return


def add_header(newsheet):
    ''' 
    This function adds header to the new sheet
    '''
    newsheet.cell(row=3, column=1).value = "轉入頂新"
    newsheet.cell(row=3, column=2).value = "單據日期"
    newsheet.cell(row=3, column=3).value = "會員編號"
    newsheet.cell(row=3, column=4).value = "託運單號"
    newsheet.cell(row=3, column=5).value = "訂單號碼"
    newsheet.cell(row=3, column=6).value = "訂單日期"
    newsheet.cell(row=3, column=7).value = "訂購人"
    newsheet.cell(row=3, column=8).value = "收貨人"
    newsheet.cell(row=3, column=9).value = "品名"
    newsheet.cell(row=3, column=10).value = "口味"
    newsheet.cell(row=3, column=11).value = "組數"
    newsheet.cell(row=3, column=12).value = "卡拉蝦"
    newsheet.cell(row=4, column=12).value = "原味"
    newsheet.cell(row=4, column=13).value = "辣味"
    newsheet.cell(row=3, column=14).value = "卡拉魷"
    newsheet.cell(row=4, column=14).value = "原味"
    newsheet.cell(row=4, column=15).value = "辣味"
    newsheet.cell(row=3, column=16).value = "卡拉蟹蟹"
    newsheet.cell(row=4, column=16).value = "芥末"
    newsheet.cell(row=4, column=17).value = "原味"
    newsheet.cell(row=4, column=18).value = "辣味"
    newsheet.cell(row=3, column=19).value = "預留"
    newsheet.cell(row=3, column=20).value = "預留"
    newsheet.cell(row=3, column=21).value = "預留"
    newsheet.cell(row=3, column=22).value = "預留"
    newsheet.cell(row=3, column=23).value = "預留"
    newsheet.cell(row=3, column=25).value = "卡拉龍珠"
    newsheet.cell(row=4, column=24).value = "原味"
    newsheet.cell(row=4, column=25).value = "辣味"
    newsheet.cell(row=4, column=26).value = "芥末"
    newsheet.cell(row=3, column=27).value = "卡拉小卷"
    newsheet.cell(row=4, column=27).value = "原味"
    newsheet.cell(row=4, column=28).value = "芥末"
    newsheet.cell(row=3, column=30).value = "虱目魚薄燒脆片"
    newsheet.cell(row=4, column=29).value = "海苔"
    newsheet.cell(row=4, column=30).value = "黑胡椒"
    newsheet.cell(row=4, column=31).value = "蒜香"
    newsheet.cell(row=3, column=32).value = "贈送"
    newsheet.cell(row=3, column=33).value = ""
    newsheet.cell(row=3, column=34).value = "備貨"
    newsheet.cell(row=3, column=35).value = "出貨"
    newsheet.cell(row=3, column=36).value = "備注"
    newsheet.cell(row=3, column=37).value = "聯絡電話"
    newsheet.cell(row=3, column=38).value = "收件人聯絡電話"
    newsheet.cell(row=3, column=39).value = "送貨地址"
    newsheet.cell(row=3, column=40).value = "訂單金額"
    newsheet.cell(row=3, column=41).value = "網路金額抵扣"
    newsheet.cell(row=3, column=42).value = "Shop Coupon"
    newsheet.cell(row=3, column=43).value = "Rakuten Coupon"
    newsheet.cell(row=3, column=44).value = "合計"
    newsheet.cell(row=3, column=45).value = "買家付款方式"
    newsheet.cell(row=3, column=46).value = "網購收款狀態"
    newsheet.cell(row=3, column=47).value = "品號"
    newsheet.cell(row=3, column=48).value = "發票"
    newsheet.cell(row=3, column=49).value = "備注"

def add_header_F(newsheet):
    ''' 
    This function adds header to the new sheet (frozen)
    '''
    newsheet.cell(row=3, column=1).value = "轉入頂新"
    newsheet.cell(row=3, column=2).value = "單據日期"
    newsheet.cell(row=3, column=3).value = "會員編號"
    newsheet.cell(row=3, column=4).value = "託運單號"
    newsheet.cell(row=3, column=5).value = "訂單號碼"
    newsheet.cell(row=3, column=6).value = "訂單日期"
    newsheet.cell(row=3, column=7).value = "訂購人"
    newsheet.cell(row=3, column=8).value = "收貨人"
    newsheet.cell(row=3, column=9).value = "商品名稱"
    newsheet.cell(row=3, column=10).value = ""
    newsheet.cell(row=3, column=11).value = "無法識別品項"
    newsheet.cell(row=3, column=12).value = "贈送"
    newsheet.cell(row=3, column=13).value = ""
    newsheet.cell(row=3, column=14).value = "備貨"
    newsheet.cell(row=3, column=15).value = "出貨"
    newsheet.cell(row=3, column=16).value = "備注"
    newsheet.cell(row=3, column=17).value = "聯絡電話"
    newsheet.cell(row=3, column=18).value = "收件人聯絡電話"
    newsheet.cell(row=3, column=19).value = "送貨地址"
    newsheet.cell(row=3, column=20).value = "訂單金額"
    newsheet.cell(row=3, column=21).value = "網路金額抵扣"
    newsheet.cell(row=3, column=22).value = "Coupon"
    newsheet.cell(row=3, column=23).value = "合計"
    newsheet.cell(row=3, column=24).value = "買家付款方式"
    newsheet.cell(row=3, column=25).value = "網購收款狀態"
    newsheet.cell(row=3, column=26).value = "品號"


def fill_row(newsheet, row, r, payment_method):
    '''
    This function fills the row from the old sheet to the new sheet.
    r is the corresponding row number in the new sheet.
    payment_method is an integer representing the method of payment
        - 13: 信用卡
        - 12: 轉帳
        - 11: 貨到付款
    '''
    # 1. date
    # 2. order date 
    newsheet.cell(row=r, column=2).value = (row[0].value[:11])
    # 3. ID number
    newsheet.cell(row=r, column=3).value = (row[55].value)
    # 4. tracking ID 
    
    # 5. order ID 
    newsheet.cell(row=r, column=5).value = (row[1].value[15:])
    # 6. order date 
    newsheet.cell(row=r, column=6).value = "20" + (row[1].value[8:14])
    # 7. customer name
    newsheet.cell(row=r, column=7).value = (row[53].value)
    # 8. receiptant name
    newsheet.cell(row=r, column=8).value = (row[61].value)
 
    # Quantity
    # 9. Product name
    newsheet.cell(row=r, column=9).value = (row[8].value.replace(" ", ""))
    # 10. flavor
    if (row[21].value != None):
        newsheet.cell(row=r, column=10).value = (row[21].value)
    elif (row[60].value != None): 
        newsheet.cell(row=r, column=10).value = (row[60].value.replace("/n", ""))
        # mark this row red to indicate remarks
        fill_color(newsheet, r, 45, 'FF0000')
    else:
        newsheet.cell(row=r, column=10).value = "無備注"
        fill_color(newsheet, r, 45, 'FF0000')
    # 11. order quantity
    newsheet.cell(row=r, column=11).value = (row[9].value)

    # 12. Kala shrimp original
    # 13. Kala shrimp spicy
    # 14. Kala you original
    # 15. Kala you spicy
    # 16. Kala you wasabi
    # 17. Kala crab original
    # 18. Kala crab spicy
    # 19. reserved
    # 20. reserved
    # 21. reserved
    # 22. reserved
    # 23. reserved
    # 24. Kala long zhu original
    # 25. Kala long zhu spicy
    # 26. Kala long zhu wasabi
    # 27. Kala xiao juan original
    # 28. Kala xiao juan wasabi
    # 29. fish chips sea weed
    # 30. fish chips black pepper
    # 31. fish chips garlic
    
    # 32. gift 1
    # 33. gift 2
    # 34. preparation
    # 35.  ship date
    newsheet.cell(row=r, column=35).value = (row[40].value)
    # 36. Note
    # 37. phone number 
    newsheet.cell(row=r, column=37).value = (row[55].value)
    # 38. contact phone number 
    newsheet.cell(row=r, column=38).value = (row[62].value)
    # 39. shipping address
    newsheet.cell(row=r, column=39).value = (row[64].value)
    # 40. order total
    newsheet.cell(row=r, column=40).value = (row[14].value)
    # 41. point spent
    newsheet.cell(row=r, column=41).value = (row[20].value)
    # 42. shop coupon
    newsheet.cell(row=r, column=42).value = (row[15].value)
    # 43. rakuten coupon
    newsheet.cell(row=r, column=43).value = (row[16].value)
    # 44. amount paid 
    newsheet.cell(row=r, column=44).value = (row[27].value)
    # 45. payment method
    newsheet.cell(row=r, column=45).value = str(payment_method) 
    # 46. payment status
    newsheet.cell(row=r, column=46).value = (row[29].value)
    # 47. code
    # 48. receipt
    if (row[59].value):
        newsheet.cell(row=r, column=48).value = (row[59].value[10:])
    # 49. note
    newsheet.cell(row=r, column=49).value = (row[60].value)

    # count quantity
    product = newsheet.cell(row=r, column=9).value
    note = newsheet.cell(row=r, column=10).value
    quantity = eval(row[9].value)
    fill_quantity(newsheet, r, product, note, quantity)

def fill_row_F(newsheet, row, r, payment_method):
    '''
    This function fills the row from the old sheet to the new sheet (frozen).
    r is the corresponding row number in the new sheet.
    payment_method is an integer representing the method of payment
        - 13: 信用卡
        - 12: 轉帳
        - 11: 貨到付款
        - 14: 點數
    '''
    # check if the order ID of the current row matches the order ID of previous row
    if (row[1].value[15:] == newsheet.cell(row=r-1, column=5).value):
        quantity = eval(row[9].value) 
        fill_quantity_F(newsheet, r-1, row[8].value, quantity)
        return r
    
    # Else, create a new row
    # 1. date
    # 2. order date 
    newsheet.cell(row=r, column=2).value = (row[0].value[:11])
    # 3. ID number
    newsheet.cell(row=r, column=3).value = (row[55].value)
    # 4. tracking ID 
    
    # 5. order ID 
    newsheet.cell(row=r, column=5).value = (row[1].value[15:])
    # 6. order date 
    newsheet.cell(row=r, column=6).value = "20" + (row[1].value[8:14])
    # 7. customer name
    newsheet.cell(row=r, column=7).value = (row[53].value)
    # 8. receiptant name
    newsheet.cell(row=r, column=8).value = (row[61].value)
    # 9. Product name
    newsheet.cell(row=r, column=9).value = ""
    newsheet.cell(row=r, column=11).value = ""
    quantity = eval(row[9].value)
    fill_quantity_F(newsheet, r, row[8].value, quantity)
    # 12. gift 1
    # 13. gift 2
    # 14. preparation
    # 15. ship date
    newsheet.cell(row=r, column=15).value = (row[40].value)
    # 16. Note
    # 17. phone number 
    newsheet.cell(row=r, column=17).value = (row[55].value)
    # 18. contact phone number 
    newsheet.cell(row=r, column=18).value = (row[62].value)
    # 19. shipping address
    newsheet.cell(row=r, column=19).value = (row[64].value)
    # 20. order total
    newsheet.cell(row=r, column=20).value = (row[14].value)
    # 21. point spent
    newsheet.cell(row=r, column=21).value = (row[20].value)
    # 22. coupon
    newsheet.cell(row=r, column=22).value = (row[15].value)
    # 23. amount paid 
    newsheet.cell(row=r, column=23).value = (row[27].value)
    # 24. payment method
    newsheet.cell(row=r, column=24).value = str(payment_method) 
    # 25. payment status
    newsheet.cell(row=r, column=25).value = (row[29].value)
    # 26. code

    # move on to next row
    return r+1  


def fill_quantity(newsheet, r, product, note, quantity):
    if (note[:2] != "口味"):
        return
    elif ("咔啦蟹蟹１包" in product):
        if ("辣" in note):
            newsheet.cell(row=r, column=18).value = str(quantity) 
        else:
            newsheet.cell(row=r, column=17).value = str(quantity)
    elif ("咔啦小卷單包" in product):
        if ("經典" in note):
            newsheet.cell(row=r, column=27).value = str(quantity)
        else:
            newsheet.cell(row=r, column=28).value = str(quantity)
    elif ("咔啦魷魚1包" in product):
        if ("原味" in note):
            newsheet.cell(row=r, column=14).value = str(quantity)
        elif ("辣" in note):
            newsheet.cell(row=r, column=15).value = str(quantity)
        else:
            newsheet.cell(row=r, column=16).value = str(quantity) 
    elif ("咔啦魷魚】1組/共3包" in product):
        if ("原味" in note):
            newsheet.cell(row=r, column=14).value = str(3 * quantity) 
        elif ("辣" in note):
            newsheet.cell(row=r, column=15).value = str(3 * quantity)
        else:
            newsheet.cell(row=r, column=16).value = str(3 * quantity)
    elif ("咔啦龍珠單包" in product):
        if ("原味" in note):
            newsheet.cell(row=r, column=24).value = str(quantity)
        elif ("辣" in note):
            newsheet.cell(row=r, column=25).value = str(quantity)
        else:
            newsheet.cell(row=r, column=26).value = str(quantity)
    elif ("咔啦蝦1包" in product):
        if ("原" in note):
            newsheet.cell(row=r, column=12).value = str(quantity)
        else:
            newsheet.cell(row=r, column=13).value = str(quantity)
    elif ("咔啦蝦】1組/共3包" in product):
        if ("原味" in note):
            newsheet.cell(row=r, column=12).value =  str(3 * quantity)
        else:
            newsheet.cell(row=r, column=13).value = str(3 * quantity)
    elif ("咔啦蝦６包" in product):
        if ("原味" in note):
            newsheet.cell(row=r, column=12).value = str(eval0(note[note.find("原味")-1]) * quantity)
        if ("辣" in note):
            newsheet.cell(row=r, column=13).value = str(eval0(note[note.find("辣")-1]) * quantity)
    elif ("咔啦魷魚６包" in product):
        if ("原味" in note):
            newsheet.cell(row=r, column=14).value = str(eval0(note[note.find("原味")-1]) * quantity)
        if ("辣味" in note):
            newsheet.cell(row=r, column=15).value = str(eval0(note[note.find("辣味")-1]) * quantity)
        if ("芥" in note):
            newsheet.cell(row=r, column=16).value = str(eval0(note[note.find("芥")-1]) * quantity)
    elif ("咔啦蟹蟹６包" in product):
        if ("原味" in note):
            newsheet.cell(row=r, column=17).value = str(eval0(note[note.find("原味")-1]) * quantity)
        if ("香蒜辣味" in note):
            newsheet.cell(row=r, column=18).value = str(eval0(note[note.find("香蒜辣味")-1]) * quantity)
    elif ("咔啦龍珠6包" in product):
        if ("原味" in note):
            newsheet.cell(row=r, column=24).value = str(eval0(note[note.find("原味")+2]) * quantity)
        if ("辣味" in note):
            newsheet.cell(row=r, column=25).value = str(eval0(note[note.find("辣味")+2]) * quantity)
        if ("芥" in note):
            newsheet.cell(row=r, column=26).value = str(eval0(note[note.find("芥")+2]) * quantity)
    elif ("虱目魚薄燒脆片70g" in product):
        if ("海苔" in note):
            newsheet.cell(row=r, column=29).value = str(quantity)
        elif ("黑胡椒" in note):
            newsheet.cell(row=r, column=30).value = str(quantity)
        else:
            newsheet.cell(row=r, column=31).value = str(quantity)
    else:
        # couldn't match the product name, mark this row red
        fill_color(newsheet, r, 49, "F8C471")


def fill_quantity_F(newsheet, r, product, quantity):
    if ("格陵蘭鴉片魚頭" in product):
        newsheet.cell(row=r, column=9).value += "格陵蘭鴉片魚頭" + "*" + str(quantity)
    elif ("熟凍冰卷 180~220g" in product):
        newsheet.cell(row=r, column=9).value += "熟凍冰卷 180~220g" + "*" + str(quantity)
    elif ("鮮蝦水餃" in product):
        newsheet.cell(row=r, column=9).value += "鮮蝦水餃" + "*" + str(quantity)
    elif ("脆皮海鮮卷" in product):
        newsheet.cell(row=r, column=9).value += "脆皮海鮮卷" + "*" + str(quantity)
    elif ("蒲燒鰻魚串" in product):
        newsheet.cell(row=r, column=9).value += "蒲燒鰻魚串" + "*" + str(quantity)
    elif ("松葉蟹味棒" in product):
        newsheet.cell(row=r, column=9).value += "松葉蟹味棒" + "*" + str(quantity)
    elif ("珍品佛跳牆" in product):
        newsheet.cell(row=r, column=9).value += "珍品佛跳牆" + "*" + str(quantity)
    elif ("薄切鯖魚片 10片/組" in product):
        newsheet.cell(row=r, column=9).value += "薄切鯖魚片 10片/組" + "*" + str(quantity)
    elif ("大比目魚薄切3入" in product):
        newsheet.cell(row=r, column=9).value += "大比目魚薄切3入" + "*" + str(quantity)
    elif ("虱目魚肚 150g" in product):
        newsheet.cell(row=r, column=9).value += "虱目魚肚 150g" + "*" + str(quantity)
    elif ("黃金花枝丸(600g/包)" in product):
        newsheet.cell(row=r, column=9).value += "黃金花枝丸(600g/包)" + "*" + str(quantity)
    elif ("鯛魚片 200-250g" in product):
        newsheet.cell(row=r, column=9).value += "鯛魚片 200-250g" + "*" + str(quantity)
    elif ("鼎級鮑魚佛跳牆" in product):
        newsheet.cell(row=r, column=9).value += "鼎級鮑魚佛跳牆" + "*" + str(quantity)
    elif ("松葉蟹鉗" in product):
        newsheet.cell(row=r, column=9).value += "松葉蟹鉗" + "*" + str(quantity)
    elif ("珍品海鮮煲" in product):
        newsheet.cell(row=r, column=9).value += "珍品海鮮煲" + "*" + str(quantity)
    elif ("櫻花蝦米糕" in product):
        newsheet.cell(row=r, column=9).value += "櫻花蝦米糕" + "*" + str(quantity)
    elif ("大白鯧魚 500-600g" in product):
        newsheet.cell(row=r, column=9).value += "大白鯧魚 500-600g" + "*" + str(quantity)
    elif ("橙汁排骨" in product):
        newsheet.cell(row=r, column=9).value += "橙汁排骨" + "*" + str(quantity)
    elif ("紅燒花筍燴元蹄" in product):
        newsheet.cell(row=r, column=9).value += "紅燒花筍燴元蹄" + "*" + str(quantity)
    elif ("麻油雞" in product):
        newsheet.cell(row=r, column=9).value += "麻油雞" + "*" + str(quantity)
    elif ("薑母鴨" in product):
        newsheet.cell(row=r, column=9).value += "薑母鴨" + "*" + str(quantity)
    elif ("鱈魚薄切3入" in product):
        newsheet.cell(row=r, column=9).value += "鱈魚薄切3入" + "*" + str(quantity)
    elif ("日式豬肉鍋" in product):
        newsheet.cell(row=r, column=9).value += "日式豬肉鍋" + "*" + str(quantity)
    elif ("冷燻鮭魚" in product):
        newsheet.cell(row=r, column=9).value += "冷燻鮭魚" + "*" + str(quantity)
    elif ("韓式泡菜鍋" in product):
        newsheet.cell(row=r, column=9).value += "韓式泡菜鍋" + "*" + str(quantity)
    elif ("虱目魚丸" in product):
        newsheet.cell(row=r, column=9).value += "虱目魚丸" + "*" + str(quantity)
    elif ("醬燒肋排" in product):
        newsheet.cell(row=r, column=9).value += "醬燒肋排" + "*" + str(quantity)
    elif ("半殼扇貝" in product):
        newsheet.cell(row=r, column=9).value += "半殼扇貝" + "*" + str(quantity)
    elif ("手工花枝漿" in product):
        newsheet.cell(row=r, column=9).value += "手工花枝漿" + "*" + str(quantity)
    elif ("櫻桃鴨胸" in product):
        newsheet.cell(row=r, column=9).value += "櫻桃鴨胸" + "*" + str(quantity)
    elif ("人蔘雞" in product):
        newsheet.cell(row=r, column=9).value += "人蔘雞" + "*" + str(quantity)
    elif ("鮮蚵卷" in product):
        newsheet.cell(row=r, column=9).value += "鮮蚵卷" + "*" + str(quantity)
    elif ("熟盤鮑 500g（8顆/包）" in product):
        newsheet.cell(row=r, column=9).value += "熟盤鮑 500g（8顆/包）" + "*" + str(quantity)
    elif ("燻燒大卷" in product):
        newsheet.cell(row=r, column=9).value += "燻燒大卷" + "*" + str(quantity)
    elif ("生吻仔魚 750g" in product):
        newsheet.cell(row=r, column=9).value += "生吻仔魚 750g" + "*" + str(quantity)
    elif ("福氣魚卵 (130克)" in product):
        newsheet.cell(row=r, column=9).value += "福氣魚卵 (130克)" + "*" + str(quantity)
    elif ("香菇貢丸" in product):
        newsheet.cell(row=r, column=9).value += "香菇貢丸" + "*" + str(quantity)
    elif ("紅豆麻糬燒" in product):
        newsheet.cell(row=r, column=9).value += "紅豆麻糬燒" + "*" + str(quantity)
    elif ("川味麻辣鍋" in product):
        newsheet.cell(row=r, column=9).value += "川味麻辣鍋" + "*" + str(quantity)
    elif ("黑蒜頭燉雞" in product):
        newsheet.cell(row=r, column=9).value += "黑蒜頭燉雞" + "*" + str(quantity)
    elif ("涼拌毛豆 150g" in product):
        newsheet.cell(row=r, column=9).value += "涼拌毛豆 150g" + "*" + str(quantity)
    elif ("老甕酸白菜鍋" in product):
        newsheet.cell(row=r, column=9).value += "老甕酸白菜鍋" + "*" + str(quantity)
    elif ("昆布味噌高湯包" in product):
        newsheet.cell(row=r, column=9).value += "昆布味噌高湯包" + "*" + str(quantity)
    elif ("油雞腿" in product):
        newsheet.cell(row=r, column=9).value += "油雞腿" + "*" + str(quantity)
    elif ("虱目魚水餃" in product):
        newsheet.cell(row=r, column=9).value += "虱目魚水餃" + "*" + str(quantity)
    elif ("人蔘烏骨雞" in product):
        newsheet.cell(row=r, column=9).value += "人蔘烏骨雞" + "*" + str(quantity)
    elif ("草蝦-12隻" in product):
        newsheet.cell(row=r, column=9).value += "草蝦12P" + "*" + str(quantity)
    elif ("蛋黃麻糬丸" in product):
        newsheet.cell(row=r, column=9).value += "蛋黃麻糬丸" + "*" + str(quantity)
    elif ("海鱺 750g" in product):
        newsheet.cell(row=r, column=9).value += "海鱺 750g" + "*" + str(quantity)
    elif ("蝦仁卷" in product):
        newsheet.cell(row=r, column=9).value += "蝦仁卷" + "*" + str(quantity)
    elif ("秋刀魚一夜干" in product):
        newsheet.cell(row=r, column=9).value += "秋刀魚一夜干" + "*" + str(quantity)
    elif ("小章魚 300g" in product):
        newsheet.cell(row=r, column=9).value += "小章魚 300g" + "*" + str(quantity)
    elif ("薄鹽鯖魚一夜干" in product):
        newsheet.cell(row=r, column=9).value += "薄鹽鯖魚一夜干" + "*" + str(quantity)
    elif ("五彩圓籠米糕" in product):
        newsheet.cell(row=r, column=9).value += "五彩圓籠米糕" + "*" + str(quantity)
    elif ("火燒蝦仁" in product):
        newsheet.cell(row=r, column=9).value += "火燒蝦仁" + "*" + str(quantity)
    elif ("熟凍海瓜子" in product):
        newsheet.cell(row=r, column=9).value += "熟凍海瓜子" + "*" + str(quantity)
    elif ("一口花枝 300g" in product):
        newsheet.cell(row=r, column=9).value += "一口花枝 300g" + "*" + str(quantity)
    elif ("烏魚子" in product and "4兩半" in product):
        newsheet.cell(row=r, column=9).value += "烏魚子4兩半" + "*" + str(quantity)
    elif ("烏魚子" in product and "4兩" in product):
        newsheet.cell(row=r, column=9).value += "烏魚子4兩" + "*" + str(quantity)
    elif ("香魚（母）每盒10尾" in product):
        newsheet.cell(row=r, column=9).value += "香魚（母）每盒10尾" + "*" + str(quantity)
    elif ("白蝦 250G/盒 (約15～16隻)" in product):
        newsheet.cell(row=r, column=9).value += "白蝦 250G/盒 (約15～16隻)" + "*" + str(quantity)
    elif ("蟹腳肉 (每盒約100g)" in product):
        newsheet.cell(row=r, column=9).value += "蟹腳肉 (每盒約100g)" + "*" + str(quantity)
    elif ("格陵蘭．純正血統鱈魚切片" in product):
        newsheet.cell(row=r, column=9).value += "格陵蘭鱈魚切片(真空包裝)" + "*" + str(quantity)
    elif ("鱈魚薄切(300gx2入)" in product):
        newsheet.cell(row=r, column=9).value += "鱈魚薄切(300gx2入)" + "*" + str(quantity)
    elif ("草蝦-8隻" in product):
        newsheet.cell(row=r, column=9).value += "草蝦8P" + "*" + str(quantity)
    elif ("板栗燒雞" in product):
        newsheet.cell(row=r, column=9).value += "板栗燒雞" + "*" + str(quantity)
    elif ("香魚（母）一尾約100g" in product):
        newsheet.cell(row=r, column=9).value += "香魚（母）一尾約100g" + "*" + str(quantity)
    elif ("薄切鯖魚片 每片約170g" in product):
        newsheet.cell(row=r, column=9).value += "薄切鯖魚片170g" + "*" + str(quantity)
    elif ("安格斯沙朗骰子牛肉 300g" in product):
        newsheet.cell(row=r, column=9).value += "安格斯沙朗骰子牛肉300g" + "*" + str(quantity)
    elif ("鯛魚下巴 (約8-10片/包) " in product):
        newsheet.cell(row=r, column=9).value += "鯛魚下巴 (約8-10片/包) " + "*" + str(quantity)
    elif ("鯰魚片 (4片，680g" in product):
        newsheet.cell(row=r, column=9).value += "鯰魚片 (4片，680g" + "*" + str(quantity)
    elif ("薄切鯖魚片 10入 (真空包裝)" in product):
        newsheet.cell(row=r, column=9).value += "薄切鯖魚片 10入 (真空包裝)" + "*" + str(quantity)
    elif ("德式帶骨香腸" in product):
        newsheet.cell(row=r, column=9).value += "德式帶骨香腸" + "*" + str(quantity)
    elif ("尖吻鱸魚" in product):
        newsheet.cell(row=r, column=9).value += "尖吻鱸魚" + "*" + str(quantity)
    elif ("多利魚片 /新鮮魴魚排 ( (4片，680g" in product):
        newsheet.cell(row=r, column=9).value += "多利魚片(4片，680g)" + "*" + str(quantity)
    elif ("珍饌櫻花蝦 100g/盒" in product):
        newsheet.cell(row=r, column=9).value += "珍饌櫻花蝦 100g/盒" + "*" + str(quantity)
    elif ("北極甜蝦 ( 250克/盒 )" in product):
        newsheet.cell(row=r, column=9).value += "北極甜蝦 ( 250克/盒 )" + "*" + str(quantity)
    elif ("虱目魚皮 (每盒300g)" in product):
        newsheet.cell(row=r, column=9).value += "虱目魚皮 (每盒300g)" + "*" + str(quantity)
    elif ("曼波魚膠 (500g" in product):
        newsheet.cell(row=r, column=9).value += "曼波魚膠 (500g)" + "*" + str(quantity)
    elif ("香魚甘露煮 200g/2入" in product):
        newsheet.cell(row=r, column=9).value += "香魚甘露煮 200g/2入" + "*" + str(quantity)
    elif ("醬漬鮭魚卵 500g" in product):
        newsheet.cell(row=r, column=9).value += "醬漬鮭魚卵 500g" + "*" + str(quantity)
    elif ("海鮮魚翅煲" in product):
        newsheet.cell(row=r, column=9).value += "海鮮魚翅煲" + "*" + str(quantity)
    elif ("黃金鱿米花 250g" in product):
        newsheet.cell(row=r, column=9).value += "黃金鱿米花 250g" + "*" + str(quantity)
    elif ("黃金蟲草燉雞" in product):
        newsheet.cell(row=r, column=9).value += "黃金蟲草燉雞" + "*" + str(quantity)
    elif ("淡菜 500g" in product):
        newsheet.cell(row=r, column=9).value += "淡菜 500g" + "*" + str(quantity)
    elif ("蒲燒鰻 500G" in product):
        newsheet.cell(row=r, column=9).value += "蒲燒鰻 500G" + "*" + str(quantity)
    elif ("美國生蠔 3入" in product):
        newsheet.cell(row=r, column=9).value += "美國生蠔 3入" + "*" + str(quantity)
    elif ("廣島牡蠣 1kg" in product):
        newsheet.cell(row=r, column=9).value += "廣島牡蠣 1kg" + "*" + str(quantity)
    elif ("丁香魚(2000g)" in product):
        newsheet.cell(row=r, column=9).value += "丁香魚(2000g)" + "*" + str(quantity)
    elif ("澎湖冰卷 140g" in product):
        newsheet.cell(row=r, column=9).value += "澎湖冰卷 140g" + "*" + str(quantity)
    elif ("日光地瓜 1000克" in product):
        newsheet.cell(row=r, column=9).value += "日光地瓜 1000克" + "*" + str(quantity)
    elif ("香螺肉 500g" in product):
        newsheet.cell(row=r, column=9).value += "香螺肉 500g" + "*" + str(quantity)
    elif ("菜圃麻糬燒" in product):
        newsheet.cell(row=r, column=9).value += "菜圃麻糬燒" + "*" + str(quantity)
    else:
        # couldn't match the product name, mark this row red
        newsheet.cell(row=r, column=11).value += product
        fill_color(newsheet, r, 27, "FF0000")

def fill_color(sheet, row, column, color):
    fillcolor = PatternFill(start_color=color, end_color=color, fill_type='solid')
    for i in range(1, column):
        sheet.cell(row=row, column=i).fill = fillcolor


# path name
# Windows
# filename = 'C:/Users/user/Desktop/樂天 版本.xlsx'
# Linux
# filename = '/home/sean/Desktop/rakuten/樂天 版本.xlsx'
filename = get_file_path()
if (filename == -1):
    exit()

# open workbook
wb = openpyxl.load_workbook(filename)

# open work sheet, by default the first sheet
sheet = wb.get_sheet_by_name('Sheet1')

# open write files
nwb1 = openpyxl.Workbook()
newfilename1 = generate_new_filename(filename, "常溫")
nwb2 = openpyxl.Workbook()
newfilename2 = generate_new_filename(filename, "冷凍")

# sheet1 信用卡
sheet1 = nwb1.active
sheet1.title = "信用卡"
add_header(sheet1)
sheet1F = nwb2.active
sheet1F.title = "信用卡"
add_header_F(sheet1F)

# sheet2 轉帳
sheet2 = nwb1.create_sheet()
sheet2.title = "轉帳"
add_header(sheet2)
sheet2F = nwb2.create_sheet()
sheet2F.title = "轉帳"
add_header_F(sheet2F)

# sheet3 貨到付款
sheet3 = nwb1.create_sheet()
sheet3.title = "貨到付款"
add_header(sheet3)
sheet3F = nwb2.create_sheet()
sheet3F.title = "貨到付款"
add_header_F(sheet3F)

# sheet4 點數
sheet4 = nwb1.create_sheet()
sheet4.title = "點數"
add_header(sheet4)
sheet4F = nwb2.create_sheet()
sheet4F.title = "點數"
add_header_F(sheet4F)

# sheet5 取消訂單
sheet5 = nwb1.create_sheet()
sheet5.title = "取消訂單"
add_header(sheet5)
sheet5F = nwb2.create_sheet()
sheet5F.title = "取消訂單"
add_header_F(sheet5F)

r1 = 5
r2 = 5
r3 = 5
r4 = 5
r1F = 5
r2F = 5
r3F = 5
r4F = 5

# process original sheet row by row
for row in sheet.iter_rows():
    if row[0].row == 1:  # skip the first line
        continue

    if ("常溫" not in row[48].value):   # 冷凍
        payment_method = sheet.cell(row=row[0].row, column=32).value
        if (payment_method == "信用卡付款"):
            r1F = fill_row_F(sheet1F, row, r1F, 13)
        elif (payment_method == "ATM轉帳"):
            r2F = fill_row_F(sheet2F, row, r2F, 12)
        elif (payment_method == "黑貓宅急便貨到付款"):
            r3F = fill_row_F(sheet3F, row, r3F, 11)
        elif (payment_method == "樂天超級點數"):
            r4F = fill_row_F(sheet4F, row, r4F, 14)
        else:
            continue

    else:                               # 常溫
        payment_method = sheet.cell(row=row[0].row, column=32).value 
        if (payment_method == "信用卡付款"):
            fill_row(sheet1, row, r1, 13)
            r1 += 1
        elif (payment_method == "ATM轉帳"):
            fill_row(sheet2, row, r2, 12)
            r2 += 1
        elif (payment_method == "黑貓宅急便貨到付款"):
            fill_row(sheet3, row, r3, 11)
            r3 += 1
        elif (payment_method == "樂天超級點數"):
            fill_row(sheet4, row, r4, 14)
            r4 += 1
        else:
            continue


# set column width
set_auto_column_widths(sheet1)
set_auto_column_widths(sheet2)
set_auto_column_widths(sheet3)
set_auto_column_widths(sheet4)
set_auto_column_widths(sheet5)
set_auto_column_widths_F(sheet1F)
set_auto_column_widths_F(sheet2F)
set_auto_column_widths_F(sheet3F)
set_auto_column_widths_F(sheet4F)
set_auto_column_widths_F(sheet5F)


# save new workbook
nwb1.save(filename = newfilename1)
nwb2.save(filename = newfilename2)

# prevent output window from closing
# input()


