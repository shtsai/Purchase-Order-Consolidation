import openpyxl
from gui import *
from openpyxl.utils import *
from openpyxl.styles import *

def set_auto_column_widths(sheet):
    '''
    This function sets the column widths automatically
    '''
    column_widths = {}
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value:
                # add 4 at the end because Chinese chars take larger spaces
                column_widths[cell.column] = max((column_widths.get(cell.column, 0), len(cell.value)+4)) 
    for i, column_width in column_widths.items():
        sheet.column_dimensions[i].width = column_width

    sheet.column_dimensions[get_column_letter(7)].width = 13
    sheet.column_dimensions[get_column_letter(8)].width = 13
    sheet.column_dimensions[get_column_letter(9)].width = 40
    sheet.column_dimensions[get_column_letter(10)].width = 20
    return

def add_header(newsheet):
    ''' 
    This function adds header to the new sheet
    '''
    newsheet.cell(row=3, column=1).value = "轉入頂新"
    newsheet.cell(row=3, column=2).value = "單據日期"
    newsheet.cell(row=3, column=3).value = "會員編號"
    newsheet.cell(row=3, column=4).value = "託運單號"
    newsheet.cell(row=3, column=5).value = "訂單號碼"
    newsheet.cell(row=3, column=6).value = "訂單日期"
    newsheet.cell(row=3, column=7).value = "訂購人"
    newsheet.cell(row=3, column=8).value = "收貨人"
    newsheet.cell(row=3, column=9).value = "品名"
    newsheet.cell(row=3, column=10).value = "數量"
    newsheet.cell(row=3, column=11).value = "卡拉蝦"
    newsheet.cell(row=4, column=11).value = "原味"
    newsheet.cell(row=4, column=12).value = "辣味"
    newsheet.cell(row=3, column=13).value = "卡拉魷"
    newsheet.cell(row=4, column=13).value = "原味"
    newsheet.cell(row=4, column=14).value = "辣味"
    newsheet.cell(row=3, column=16).value = "卡拉蟹蟹"
    newsheet.cell(row=4, column=15).value = "芥末"
    newsheet.cell(row=4, column=16).value = "原味"
    newsheet.cell(row=4, column=17).value = "辣味"
    newsheet.cell(row=3, column=18).value = "預留"
    newsheet.cell(row=3, column=19).value = "預留"
    newsheet.cell(row=3, column=20).value = "預留"
    newsheet.cell(row=3, column=21).value = "預留"
    newsheet.cell(row=3, column=22).value = "預留"
    newsheet.cell(row=3, column=24).value = "卡拉龍珠"
    newsheet.cell(row=4, column=23).value = "原味"
    newsheet.cell(row=4, column=24).value = "辣味"
    newsheet.cell(row=4, column=25).value = "芥末"
    newsheet.cell(row=3, column=26).value = "卡拉小卷"
    newsheet.cell(row=4, column=26).value = "原味"
    newsheet.cell(row=4, column=27).value = "芥末"
    newsheet.cell(row=3, column=29).value = "虱目魚薄燒脆片"
    newsheet.cell(row=4, column=28).value = "海苔"
    newsheet.cell(row=4, column=29).value = "黑胡椒"
    newsheet.cell(row=4, column=30).value = "蒜香"
    newsheet.cell(row=3, column=31).value = "贈送"
    newsheet.cell(row=3, column=32).value = ""
    newsheet.cell(row=3, column=33).value = "備貨"
    newsheet.cell(row=3, column=34).value = "出貨"
    newsheet.cell(row=3, column=35).value = "備注"
    newsheet.cell(row=3, column=36).value = "聯絡電話"
    newsheet.cell(row=3, column=37).value = "收件人聯絡電話"
    newsheet.cell(row=3, column=38).value = "送貨地址"
    newsheet.cell(row=3, column=39).value = "訂單金額"
    newsheet.cell(row=3, column=40).value = "網路金額抵扣"
    newsheet.cell(row=3, column=41).value = "Coupon"
    newsheet.cell(row=3, column=42).value = "合計"
    newsheet.cell(row=3, column=43).value = "買家付款方式"
    newsheet.cell(row=3, column=44).value = "網購收款狀態"
    newsheet.cell(row=3, column=45).value = "品號"

def fill_row(newsheet, row, r, payment_method):
    '''
    This function fills the row from the old sheet to the new sheet.
    r is the corresponding row number in the new sheet.
    payment_method is an integer representing the method of payment
        - 13: 信用卡
        - 12: 轉帳
        - 11: 貨到付款
    '''
    # 1. date
    # 2. order date 
    newsheet.cell(row=r, column=2).value = (row[0].value[:11])
    # 3. ID number
    newsheet.cell(row=r, column=3).value = (row[55].value)
    # 4. tracking ID 
    
    # 5. order ID 
    newsheet.cell(row=r, column=5).value = (row[1].value[15:])
    # 6. order date 
    newsheet.cell(row=r, column=6).value = "20" + (row[1].value[8:14])
    # 7. customer name
    newsheet.cell(row=r, column=7).value = (row[53].value)
    # 8. receiptant name
    newsheet.cell(row=r, column=8).value = (row[61].value)
 
    # Quantity
    # 9. Product name
    newsheet.cell(row=r, column=9).value = (row[8].value.replace(" ", "")[:25])
    # 10. quantity
    if (row[21].value != None):
        newsheet.cell(row=r, column=10).value = (row[21].value)
    elif (row[60].value != None): 
        newsheet.cell(row=r, column=10).value = (row[60].value.replace("/n", ""))
        # mark this row to indicate remarks
        red_fill(newsheet, r)
    else:
        newsheet.cell(row=r, column=10).value = "無備注"

    # 11. Kala shrimp original
    # 12. Kala shrimp spicy
    # 13. Kala you original
    # 14. Kala you spicy
    # 15. Kala you wasabi
    # 16. Kala crab original
    # 17. Kala crab spicy
    # 18. reserved
    # 19. reserved
    # 20. reserved
    # 21. reserved
    # 22. reserved
    # 23. Kala long zhu original
    # 24. Kala long zhu spicy
    # 25. Kala long zhu wasabi
    # 26. Kala xiao juan original
    # 27. Kala xiao juan wasabi
    # 28. fish chips sea weed
    # 29. fish chips black pepper
    # 30. fish chips garlic
    
    # 31. gift 1
    # 32. gift 2
    # 33. preparation
    # 34. ship date
    newsheet.cell(row=r, column=34).value = (row[40].value)
    # 35. Note
    # 36. phone number 
    newsheet.cell(row=r, column=36).value = (row[55].value)
    # 37. contact phone number 
    newsheet.cell(row=r, column=37).value = (row[62].value)
    # 38. shipping address
    newsheet.cell(row=r, column=38).value = (row[64].value)
    # 39. order total
    newsheet.cell(row=r, column=39).value = (row[14].value)
    # 40. point spent
    newsheet.cell(row=r, column=40).value = (row[20].value)
    # 41. coupon
    newsheet.cell(row=r, column=41).value = (row[15].value)
    # 42. amount paid 
    newsheet.cell(row=r, column=42).value = (row[27].value)
    # 43. payment method
    newsheet.cell(row=r, column=43).value = str(payment_method) 
    # 44. payment status
    newsheet.cell(row=r, column=44).value = (row[29].value)
    # 45. code
    
    # count quantity
    product = newsheet.cell(row=r, column=9).value
    remark = newsheet.cell(row=r, column=10).value
    if ("咔啦蟹蟹1包" in product):
        if ("辣" in product):
            newsheet.cell(row=r, column=16).value = "1"
        else:
            newsheet.cell(row=r, column=17).value = "1"


def red_fill(sheet, row):
    redfill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
    for i in range(1, 45):
        sheet.cell(row=row, column=i).fill = redfill


# path name
# Windows
# filename = 'C:/Users/user/Desktop/樂天 版本.xlsx'
# Linux
# filename = '/home/sean/Desktop/rakuten/樂天 版本.xlsx'
filename = get_file_path()
if (filename == -1):
    exit()

# open workbook
wb = openpyxl.load_workbook(filename)

# open work sheet, by default the first sheet
sheet = wb.get_sheet_by_name('Sheet1')

# open write files
nwb1 = openpyxl.Workbook()
newfilename1 = generate_new_filename(filename, "常溫")

# sheet1 信用卡
sheet1 = nwb1.active
sheet1.title = "信用卡"
add_header(sheet1)

# sheet2 轉帳
sheet2 = nwb1.create_sheet()
sheet2.title = "轉帳"
add_header(sheet2)

# sheet3 貨到付款
sheet3 = nwb1.create_sheet()
sheet3.title = "貨到付款"
add_header(sheet3)

# sheet4 點數
sheet4 = nwb1.create_sheet()
sheet4.title = "點數"
add_header(sheet4)

# sheet5 取消訂單
sheet5 = nwb1.create_sheet()
sheet5.title = "取消訂單"
add_header(sheet5)

r1 = 5
r2 = 5
r3 = 5
r4 = 5
# print all non-empty cells
for row in sheet.iter_rows():
    if row[0].row == 1:  # skip the first line
        continue

    if ("常溫" not in row[48].value):   # only handle 常溫 for now
        continue
    payment_method = sheet.cell(row=row[0].row, column=32).value 
    if (payment_method == "信用卡付款"):
        fill_row(sheet1, row, r1, 13)
        r1 += 1
    elif (payment_method == "ATM轉帳"):
        fill_row(sheet2, row, r2, 12)
        r2 += 1
    elif (payment_method == "黑貓宅急便貨到付款"):
        fill_row(sheet3, row, r3, 11)
        r3 += 1
    elif (payment_method == "樂天超級點數"):
        fill_row(sheet4, row, r4, 14)
    else:
        continue


# set column width
set_auto_column_widths(sheet1)
set_auto_column_widths(sheet2)
set_auto_column_widths(sheet3)
set_auto_column_widths(sheet4)
set_auto_column_widths(sheet5)


# save new workbook
nwb1.save(filename = newfilename1)


# prevent output window from closing
# input()


