import openpyxl
from gui import *
from number_parser import *
from openpyxl.utils import *
from openpyxl.styles import *
from openpyxl.styles.borders import Border, Side

def set_auto_column_widths(sheet):
    '''
    This function sets the column widths automatically
    '''
    column_widths = {}
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value:
                # add 4 at the end because Chinese chars take larger spaces
                column_widths[cell.column] = max((column_widths.get(cell.column, 0), len(cell.value)+4)) 
    for i, column_width in column_widths.items():
        sheet.column_dimensions[i].width = column_width

    sheet.column_dimensions[get_column_letter(7)].width = 13
    sheet.column_dimensions[get_column_letter(8)].width = 13
    sheet.column_dimensions[get_column_letter(9)].width = 40
    sheet.column_dimensions[get_column_letter(10)].width = 20
    return

def add_header(newsheet):
    ''' 
    This function adds header to the new sheet (frozen)
    '''
    newsheet.cell(row=3, column=1).value = "轉入頂新"
    newsheet.cell(row=3, column=2).value = "單據日期"
    newsheet.cell(row=3, column=3).value = "會員編號"
    newsheet.cell(row=3, column=4).value = "託運單號"
    newsheet.cell(row=3, column=5).value = "訂單號碼"
    newsheet.cell(row=3, column=6).value = "訂單日期"
    newsheet.cell(row=3, column=7).value = "訂購人"
    newsheet.cell(row=3, column=8).value = "收貨人"
    newsheet.cell(row=3, column=9).value = "商品名稱"
    newsheet.cell(row=3, column=10).value = "無法識別品項"
    newsheet.cell(row=3, column=11).value = "組數"
    newsheet.cell(row=3, column=12).value = "贈送"
    newsheet.cell(row=3, column=13).value = ""
    newsheet.cell(row=3, column=14).value = "備貨"
    newsheet.cell(row=3, column=15).value = "出貨"
    newsheet.cell(row=3, column=16).value = "備注"
    newsheet.cell(row=3, column=17).value = "聯絡電話"
    newsheet.cell(row=3, column=18).value = "收件人聯絡電話"
    newsheet.cell(row=3, column=19).value = "送貨地址"
    newsheet.cell(row=3, column=20).value = "訂單金額"
    newsheet.cell(row=3, column=21).value = "超贈點點數"
    newsheet.cell(row=3, column=22).value = "合計"
    newsheet.cell(row=3, column=23).value = ""
    newsheet.cell(row=3, column=24).value = ""
    newsheet.cell(row=3, column=25).value = "買家付款方式"
    newsheet.cell(row=3, column=26).value = "網購收款狀態"
    newsheet.cell(row=3, column=27).value = "品號"

def fill_row(newsheet, row, r, payment_method):
    '''
    This function fills the row from the old sheet to the new sheet (frozen).
    r is the corresponding row number in the new sheet.
    payment_method is an integer representing the method of payment
        - 13: 信用卡
        - 12: 轉帳
    '''
    # check if the order ID of the current row matches the order ID of previous row
    # if they don't match, add a border to separate them
    '''
    if (row[1].value[15:] != newsheet.cell(row=r-1, column=5).value):
        border = Border(top=Side(style="thick", color="1E90FF"))  #<--change border color here
        for i in range(1, 30):
            newsheet.cell(row=r, column=i).border = border
    '''

    # 1. date
    # 2. order date 
    newsheet.cell(row=r, column=2).value = (row[0].value[:11])
    # 3. ID number
    newsheet.cell(row=r, column=3).value = (row[55].value)
    # 4. tracking ID 
    
    # 5. order ID 
    newsheet.cell(row=r, column=5).value = (row[1].value[15:])
    # 6. order date 
    newsheet.cell(row=r, column=6).value = "20" + (row[1].value[8:14])
    # 7. customer name
    newsheet.cell(row=r, column=7).value = (row[53].value)
    # 8. receiptant name
    newsheet.cell(row=r, column=8).value = (row[61].value)
    # 9. Product name
    newsheet.cell(row=r, column=9).value = ""
    newsheet.cell(row=r, column=10).value = ""
    quantity = eval(row[9].value)
    fill_quantity_F(newsheet, r, row[8].value, quantity)
    # 10. unrecognizable product
    # 11. quantity
    newsheet.cell(row=r, column=11).value = (row[9].value)
    # 12. gift 1
    # 13. gift 2
    # 14. preparation
    # 15. ship date
    newsheet.cell(row=r, column=15).value = (row[40].value)
    # 16. Note
    # 17. phone number 
    newsheet.cell(row=r, column=17).value = (row[55].value)
    # 18. contact phone number 
    newsheet.cell(row=r, column=18).value = (row[62].value)
    # 19. shipping address
    newsheet.cell(row=r, column=19).value = (row[64].value)
    # 20. order total
    newsheet.cell(row=r, column=20).value = (row[14].value)
    # 21. point spent
    newsheet.cell(row=r, column=21).value = (row[20].value)
    # 22. Shop Coupon
    newsheet.cell(row=r, column=22).value = (row[15].value)
    # 23. Rakuten Coupon
    newsheet.cell(row=r, column=23).value = (row[16].value)
    # 24. amount paid 
    newsheet.cell(row=r, column=24).value = (row[27].value)
    # 25. payment method
    newsheet.cell(row=r, column=25).value = str(payment_method) 
    # 26. payment status
    newsheet.cell(row=r, column=26).value = (row[29].value)
    # 27. code
    # 28. receipt
    if (row[59].value):
        newsheet.cell(row=r, column=28).value = (row[59].value[10:])
    # 29. note
    newsheet.cell(row=r, column=29).value = (row[60].value)

    # move on to next row
    return r+1  


def fill_color(sheet, row, column, color):
    fillcolor = PatternFill(start_color=color, end_color=color, fill_type='solid')
    for i in range(1, column):
        sheet.cell(row=row, column=i).fill = fillcolor


# path name
filename = get_file_path()
if (filename == -1):
    exit()

# open workbook
wb = openpyxl.load_workbook(filename)

# open work sheet, by default the first sheet
sheet = wb.get_sheet_by_name(wb.get_sheet_name()[0])

# open write files
nwb1 = openpyxl.Workbook()
newfilename1 = generate_new_filename(filename, "冷凍")

# sheet1 信用卡
sheet1 = nwb1.active
sheet1.title = "信用卡"
add_header(sheet1)

# sheet2 轉帳
sheet2 = nwb1.create_sheet()
sheet2.title = "轉帳"
add_header(sheet2)

# sheet3 取消訂單
sheet3 = nwb1.create_sheet()
sheet3.title = "取消訂單"
add_header(sheet3)

r1 = 5
r2 = 5

# process original sheet row by row
for row in sheet.iter_rows():
    if row[0].row == 1:  # skip the first line
        continue

    payment_method = sheet.cell(row=row[0].row, column=2).value
    if ("信用卡" in payment_method):
        r1F = fill_row_F(sheet1, row, r1, 13)
    elif ("ATM" in payment_method):
        r2F = fill_row_F(sheet2, row, r2, 12)
    else:
        continue

# set column width
set_auto_column_widths(sheet1)
set_auto_column_widths(sheet2)
set_auto_column_widths(sheet3)


# save new workbook
nwb1.save(filename = newfilename1)

# prevent output window from closing
# input()


