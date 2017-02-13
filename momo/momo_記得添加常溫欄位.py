import openpyxl
from gui import *
from openpyxl.utils import *
from openpyxl.styles import *
from openpyxl.styles.borders import Border, Side

import xlrd

def set_auto_column_widths(sheet):
    '''
    This function sets the column widths automatically
    '''
    column_widths = {}
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value:
                # add 4 at the end because Chinese chars take larger spaces
                column_widths[cell.column] = max((column_widths.get(cell.column, 0), len(cell.value)+4)) 
    for i, column_width in column_widths.items():
        sheet.column_dimensions[i].width = column_width

    
    sheet.column_dimensions[get_column_letter(9)].width = 40
    return


def add_header(newsheet):
    ''' 
    This function adds header to the new sheet (frozen)
    '''
    newsheet.cell(row=3, column=1).value = "轉入頂新"
    newsheet.cell(row=3, column=2).value = "單據日期"
    newsheet.cell(row=3, column=3).value = "會員編號"
    newsheet.cell(row=3, column=4).value = "託運單號"
    newsheet.cell(row=3, column=5).value = "訂單號碼"
    newsheet.cell(row=3, column=6).value = "訂單日期"
    newsheet.cell(row=3, column=7).value = "訂購人"
    newsheet.cell(row=3, column=8).value = "收貨人"
    newsheet.cell(row=3, column=9).value = "商品名稱"
    newsheet.cell(row=3, column=10).value = ""
    newsheet.cell(row=3, column=11).value = "組數"
    newsheet.cell(row=3, column=12).value = "贈送"
    newsheet.cell(row=3, column=13).value = ""
    newsheet.cell(row=3, column=14).value = "備貨"
    newsheet.cell(row=3, column=15).value = "出貨"
    newsheet.cell(row=3, column=16).value = "備注"
    newsheet.cell(row=3, column=17).value = "行動電話"
    newsheet.cell(row=3, column=18).value = "電話"
    newsheet.cell(row=3, column=19).value = "送貨地址"
    newsheet.cell(row=3, column=20).value = "訂單金額"
    newsheet.cell(row=3, column=21).value = "合計"
    newsheet.cell(row=3, column=22).value = ""
    newsheet.cell(row=3, column=23).value = ""
    newsheet.cell(row=3, column=24).value = ""
    newsheet.cell(row=3, column=25).value = "買家付款方式"
    newsheet.cell(row=3, column=26).value = "網購收款狀態"
    newsheet.cell(row=3, column=27).value = "品號"


def fill_row(newsheet, row, r, payment_method):
    '''
    This function fills the row from the old sheet to the new sheet (frozen).
    r is the corresponding row number in the new sheet.
    payment_method is an integer representing the method of payment
    '''
    # check if the order ID of the current row matches the order ID of previous row
    # if they don't match, add a border to separate them
    
    if (row[1].value[:14] != newsheet.cell(row=r-1, column=5).value):
        border = Border(top=Side(style="thick", color="1E90FF"))  #<--change border color here
        for i in range(1, 30):
            newsheet.cell(row=r, column=i).border = border
    
    # 1. date
    # 2. order date
    newsheet.cell(row=r, column=2).value = str(row[7].value[:10])
    # 3. member ID number
    phone = str(row[11].value)
    if phone[0] != '0':
        phone = '0' + phone
    phone = phone.split(".")[0]
    newsheet.cell(row=r, column=3).value = phone
    # 4. tracking ID 
    
    # 5. order ID 
    newsheet.cell(row=r, column=5).value = str(row[1].value[:14])
    # 6. order date 
    newsheet.cell(row=r, column=6).value = str(row[7].value[:10])
    # 7. customer name
    newsheet.cell(row=r, column=7).value = str(row[9].value)
    # 8. receiptant name
    newsheet.cell(row=r, column=8).value = str(row[9].value)
    # 9. Product name
    newsheet.cell(row=r, column=9).value = str(row[15].value).replace("【海鮮主義】","")
    # 10. unrecognizable product#    
    newsheet.cell(row=r, column=10).value = ""
    # 11. quantity
    newsheet.cell(row=r, column=11).value = str(row[17].value)
    # 12. gift 1
    # 13. gift 2
    # 14. preparation
    # 15. ship date
    newsheet.cell(row=r, column=15).value = str(row[8].value)
    # 16. Note
    # 17. cellphone number 
    newsheet.cell(row=r, column=17).value = phone
    # 18. phone number 
    newsheet.cell(row=r, column=18).value = str(row[10].value)
    # 19. shipping address
    newsheet.cell(row=r, column=19).value = str(row[12].value)
    # 20. order amount
    newsheet.cell(row=r, column=20).value = str(row[18].value)
    # 21. total
    # 22. reserved
    # 23. reserved
    # 24. reserved
    # 25. payment method
    newsheet.cell(row=r, column=25).value = payment_method
    # 26. payment status
    newsheet.cell(row=r, column=26).value = "已付款"
    # 27. code

    # move on to next row
    return r+1  

def calculate_total(sheet):
    '''
    This function sums the amounts for each order and calculates total ammount.
    '''
    start = 5
    total = 0
    for r in range(5, sheet.max_row+1):
        if (sheet.cell(row=r, column=5).value == sheet.cell(row=r+1, column=5).value):
            total += eval(sheet.cell(row=r,column=20).value)
            continue
        total += eval(sheet.cell(row=r,column=20).value)
        for i in range(start, r+1):
            sheet.cell(row=i, column=21).value = str(total)
        total = 0
        start = r+1

def fill_color(sheet, row, column, color):
    fillcolor = PatternFill(start_color=color, end_color=color, fill_type='solid')
    for i in range(1, column):
        sheet.cell(row=row, column=i).fill = fillcolor


# path name
filename = get_file_path()
if (filename == -1):
    exit()

# open workbook
wb = xlrd.open_workbook(filename)

# open work sheet, by default the first sheet
sheet = wb.sheet_by_index(0)

# open write files
nwb1 = openpyxl.Workbook()
newfilename1 = generate_new_filename(filename, "常溫")
nwb2 = openpyxl.Workbook()
newfilename2 = generate_new_filename(filename, "冷凍")

# sheet1 信用卡
sheet1 = nwb1.active
sheet1.title = "信用卡"
add_header(sheet1)
sheet1F = nwb2.active
sheet1F.title = "信用卡"
add_header(sheet1F)

# sheet2 轉帳
sheet2 = nwb1.create_sheet()
sheet2.title = "轉帳"
add_header(sheet2)
sheet2F = nwb2.create_sheet()
sheet2F.title = "轉帳"
add_header(sheet2F)

# sheet3 貨到付款
sheet3 = nwb1.create_sheet()
sheet3.title = "取消訂單"
add_header(sheet3)
sheet3F = nwb2.create_sheet()
sheet3F.title = "取消訂單"
add_header(sheet3F)


r1 = 5
r2 = 5
r1F = 5
r2F = 5

# process original sheet row by row
for r in range(1, sheet.nrows):
    row = sheet.row(r)

    if ("常溫" not in row[25].value):   # 冷凍
        payment_method = row[19].value
        if (payment_method == "信用卡"):
            r1F = fill_row(sheet1F, row, r1F, "13")
        elif (payment_method == "ATM"):
            r2F = fill_row(sheet2F, row, r2F, "12")
        else:
            continue

    else:                               # 常溫
        payment_method = row[19].value 
        if (payment_method == "信用卡"):
            fill_row(sheet1, row, r1, "13")
            r1 += 1
        elif (payment_method == "ATM"):
            fill_row(sheet2, row, r2, "12")
            r2 += 1
        else:
            continue

# calculate total
calculate_total(sheet1F)
calculate_total(sheet2F)
calculate_total(sheet1)
calculate_total(sheet2)

# set column width
set_auto_column_widths(sheet1)
set_auto_column_widths(sheet2)
set_auto_column_widths(sheet3)
set_auto_column_widths(sheet1F)
set_auto_column_widths(sheet2F)
set_auto_column_widths(sheet3F)


# save new workbook
nwb1.save(filename = newfilename1)
nwb2.save(filename = newfilename2)

# prevent output window from closing
# input()


