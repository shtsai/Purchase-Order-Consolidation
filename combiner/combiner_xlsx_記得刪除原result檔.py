import openpyxl
from openpyxl.utils import *
from openpyxl.styles import *
from openpyxl.styles.borders import Border, Side

import tkinter
from tkinter import filedialog
import os

def get_file_path():
    root = tkinter.Tk()
    root.withdraw()   # hide the gui form

    # open up a file dialog to obtain the file path
    file_path = filedialog.askopenfilenames(filetypes=[('Microsoft Excel Worksheet','*.xlsx')])

    return file_path


def set_auto_column_widths(sheet):
    '''
    This function sets the column widths automatically
    '''
    column_widths = {}
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value:
                # add 4 at the end because Chinese chars take larger spaces
                column_widths[cell.column] = max((column_widths.get(cell.column, 0), len(cell.value)+4)) 
    for i, column_width in column_widths.items():
        sheet.column_dimensions[i].width = column_width

    return

def fill_row(newsheet, row, r):
    for i in range(len(row)):
        if (type(row[i].value) == float):
            newsheet.cell(row=r, column=i+1).value = str(int(row[i].value))
        else:
            newsheet.cell(row=r, column=i+1).value = str(row[i].value)


# path name
filename = get_file_path()

# open write files
nwb = openpyxl.Workbook()
newfilename = "result.xlsx"

sheet1 = nwb.active
sheet1.title = "信用卡"

r = 2
for i in range(len(filename)):
   # open workbook
    wb = openpyxl.load_workbook(filename[i])

    # open work sheet, by default the first sheet
    sheet = wb.get_sheet_by_name(wb.get_sheet_names()[0])
 
    # process original sheet row by row
    for row in sheet.iter_rows():
        if (i == 0 and row[0].row == 1):
            fill_row(sheet1, row, 1)
        if row[0].row == 1:  # skip the first line
            continue

        fill_row(sheet1, row, r)
        r += 1

# set column width
set_auto_column_widths(sheet1)

# save new workbook
nwb.save(filename = newfilename)

# prevent output window from closing
# input()


