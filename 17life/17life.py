import openpyxl
from gui import *
from openpyxl.utils import *
from openpyxl.styles import *
from openpyxl.styles.borders import Border, Side

import xlrd

def set_auto_column_widths(sheet):
    '''
    This function sets the column widths automatically
    '''
    column_widths = {}
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value:
                # add 4 at the end because Chinese chars take larger spaces
                column_widths[cell.column] = max((column_widths.get(cell.column, 0), len(cell.value)+4)) 
    for i, column_width in column_widths.items():
        sheet.column_dimensions[i].width = column_width

    
    sheet.column_dimensions[get_column_letter(9)].width = 40
    return


def add_header(newsheet):
    ''' 
    This function adds header to the new sheet (frozen)
    '''
    newsheet.cell(row=3, column=1).value = "轉入頂新"
    newsheet.cell(row=3, column=2).value = "單據日期"
    newsheet.cell(row=3, column=3).value = "會員編號"
    newsheet.cell(row=3, column=4).value = "託運單號"
    newsheet.cell(row=3, column=5).value = "訂單號碼"
    newsheet.cell(row=3, column=6).value = "訂單日期"
    newsheet.cell(row=3, column=7).value = "訂購人"
    newsheet.cell(row=3, column=8).value = "收貨人"
    newsheet.cell(row=3, column=9).value = "商品名稱"
    newsheet.cell(row=3, column=10).value = ""
    newsheet.cell(row=3, column=11).value = ""
    newsheet.cell(row=3, column=12).value = "贈送"
    newsheet.cell(row=3, column=13).value = ""
    newsheet.cell(row=3, column=14).value = "備貨"
    newsheet.cell(row=3, column=15).value = "出貨"
    newsheet.cell(row=3, column=16).value = "備注"
    newsheet.cell(row=3, column=17).value = "行動電話"
    newsheet.cell(row=3, column=18).value = "電話"
    newsheet.cell(row=3, column=19).value = "送貨地址"
    newsheet.cell(row=3, column=20).value = "訂單金額"
    newsheet.cell(row=3, column=21).value = ""
    newsheet.cell(row=3, column=22).value = ""
    newsheet.cell(row=3, column=23).value = ""
    newsheet.cell(row=3, column=24).value = "合計"
    newsheet.cell(row=3, column=25).value = "買家付款方式"
    newsheet.cell(row=3, column=26).value = "網購收款狀態"
    newsheet.cell(row=3, column=27).value = "品號"


def fill_row(newsheet, row, r, payment_method):
    '''
    This function fills the row from the old sheet to the new sheet (frozen).
    r is the corresponding row number in the new sheet.
    '''
    # check if the order ID of the current row matches the order ID of previous row
    # if they don't match, add a border to separate them
    ''' 
    if (row[1].value[:14] != newsheet.cell(row=r-1, column=5).value):
        border = Border(top=Side(style="thick", color="1E90FF"))  #<--change border color here
        for i in range(1, 30):
            newsheet.cell(row=r, column=i).border = border
    '''

    # 1. date
    # 2. order date
    newsheet.cell(row=r, column=2).value = str(row[2].value[:10])
    # 3. member ID number
    newsheet.cell(row=r, column=3).value = str(row[4].value)
    # 4. tracking ID 
    
    # 5. order ID 
    newsheet.cell(row=r, column=5).value = str(row[1].value[:16]).replace("-","")
    # 6. order date 
    newsheet.cell(row=r, column=6).value = str(row[2].value[:10])
    # 7. customer name
    newsheet.cell(row=r, column=7).value = str(row[3].value)
    # 8. receiptant name
    newsheet.cell(row=r, column=8).value = str(row[3].value)
    # 9. Product name
    newsheet.cell(row=r, column=9).value = ""
    products = str(row[10].value).replace("【饗城】","").replace("【海鮮主義】","").split("\n")
    quantities = str(row[11].value).split("\n")
    for i in range(len(products)):
        newsheet.cell(row=r, column=9).value += products[i] + "*" + quantities[i] + " " 
    # 10. unrecognizable product    
    newsheet.cell(row=r, column=10).value = ""
    # 11. quantity
    newsheet.cell(row=r, column=11).value = "" 
    # 12. gift 1
    # 13. gift 2
    # 14. preparation
    # 15. ship date
#    newsheet.cell(row=r, column=15).value = ""
    # 16. Note
    # 17. cellphone number 
    newsheet.cell(row=r, column=17).value = str(row[4].value) 
    # 18. phone number 
    newsheet.cell(row=r, column=18).value = str(row[4].value)
    # 19. shipping address
    newsheet.cell(row=r, column=19).value = str(row[5].value)
    # 20. order amount
#    newsheet.cell(row=r, column=20).value = ""
    # 21. reserved
    # 22. reserved
    # 23.reserved
    # 24. total amount
#    newsheet.cell(row=r, column=24).value = ""
    # 25. payment method
#    newsheet.cell(row=r, column=25).value = ""
    # 26. payment status
#    newsheet.cell(row=r, column=26).value = ""
    # 27. code

    # move on to next row
    return r+1  

def fill_color(sheet, row, column, color):
    fillcolor = PatternFill(start_color=color, end_color=color, fill_type='solid')
    for i in range(1, column):
        sheet.cell(row=row, column=i).fill = fillcolor


# path name
filename = get_file_path()
if (filename == -1):
    exit()

# open workbook
wb = xlrd.open_workbook(filename)

# open work sheet, by default the first sheet
sheet = wb.sheet_by_index(0)

# open write files
nwb1 = openpyxl.Workbook()
newfilename1 = generate_new_filename(filename, "冷凍")

# sheet1 信用卡
sheet1 = nwb1.active
sheet1.title = "信用卡"
add_header(sheet1)

# sheet2 取消訂單 
sheet2 = nwb1.create_sheet()
sheet2.title = "取消訂單"
add_header(sheet2)


r1 = 5

# process original sheet row by row
for r in range(1, sheet.nrows):
    row = sheet.row(r)

    r1 = fill_row(sheet1, row, r1, "信用卡")


# set column width
set_auto_column_widths(sheet1)
set_auto_column_widths(sheet2)

# save new workbook
nwb1.save(filename = newfilename1)

# prevent output window from closing
# input()


